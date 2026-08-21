import os
import sys
import time
import argparse
import logging
from typing import List, Optional
from datetime import datetime
from pathlib import Path

from config import Config
from tb_client import ThingsBoardClient
from report_generator import generate_html_report
from mail_sender import EmailSender

logger = logging.getLogger("BBMP_Panel_Report.Main")


def get_mock_panel_data(project_name: str = "BBMP"):
    """Generates realistic mock installation report data for testing and preview purposes."""
    is_bbmp = (project_name.upper() == "BBMP")
    proj = "BBMP" if is_bbmp else "5B Innovation"
    
    regions = [
        {"name": f"{proj} North Region", "online": 1250, "offline": 15, "offline_pf_today": 10, "offline_pf_prior": 15, "offline_pf": 25, "total": 1290},
        {"name": f"{proj} Central Region", "online": 1820, "offline": 20, "offline_pf_today": 12, "offline_pf_prior": 18, "offline_pf": 30, "total": 1870},
        {"name": f"{proj} South Region", "online": 905, "offline": 13, "offline_pf_today": 8, "offline_pf_prior": 12, "offline_pf": 20, "total": 938}
    ] if not is_bbmp else []

    return {
        "installation_report": {
            "project_name": proj,
            "performance_score": 98.83,
            "kpi_score": 94.92,
            "penalty_points": 208,
            "offline_gt_7_days": 18,
            "offline_pf_gt_7_days": 24,
            "total_panels": 4098,
            "regions": regions,
            "bommanahalli": {"online": 1408, "offline": 14, "offline_pf_today": 12, "offline_pf_prior": 24, "offline_pf": 36, "total": 1458},
            "east": {"online": 2567, "offline": 34, "offline_pf_today": 15, "offline_pf_prior": 24, "offline_pf": 39, "total": 2640},
            "combined": {"online": 3975, "offline": 48, "offline_pf_today": 27, "offline_pf_prior": 48, "offline_pf": 75, "total": 4098},
            "issues": {
                "low_voltage": 7, "high_voltage": 3, "power_failure": 179,
                "high_current": 5, "low_current": 7, "mcb_trip": 22,
                "relay_failure": 0, "meter_comm_failure": 0, "manual_operation": 0, "panel_door_open": 21,
                "offline_gt_7_days": 18, "offline_pf_gt_7_days": 24
            },
            "affected_panels": [
                {"id": "dev-101", "name": f"Panel-{proj}-001", "label": "Main Sector 1", "region": "NORTH", "zone": "Zone A", "ward": "Ward 10", "status": "ONLINE", "last_received_date": "2026-08-18 11:59:45", "days_offline": "-", "active_issues": ["Low Voltage"], "lat_lon": "12.97391, 77.64478"},
                {"id": "dev-102", "name": f"Panel-{proj}-002", "label": "Central Sector 2", "region": "SOUTH", "zone": "Zone B", "ward": "Ward 12", "status": "OFFLINE", "last_received_date": "2026-08-10 10:30:00", "days_offline": "8 Days", "active_issues": ["High Voltage", "Offline (>7 Days)"], "lat_lon": "13.01162, 77.61090"},
                {"id": "dev-103", "name": f"Panel-{proj}-005", "label": "East Highway 5", "region": "EAST", "zone": "Zone C", "ward": "Ward 15", "status": "OFFLINE_PF_PRIOR", "last_received_date": "2026-08-06 14:15:20", "days_offline": "12 Days", "active_issues": ["High Current / Power Theft", "Offline PF (>7 Days)"], "lat_lon": "13.00363, 77.62065"}
            ]
        },
        "summary": {"total_devices": 4098, "online_devices": 3975, "offline_devices": 123, "active_alarms": 1},
        "alarms": [
            {
                "type": "LOW_VOLTAGE_ALARM",
                "severity": "CRITICAL",
                "originatorName": f"Panel-{proj}-001",
                "createdTime": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
        ]
    }


import sys
import io

# Force UTF-8 output encoding for Windows terminals
if hasattr(sys.stdout, 'buffer'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')


def print_terminal_summary(data):
    inst = data.get("installation_report", {})
    proj_name = inst.get("project_name", "BBMP")
    regions = inst.get("regions", [])
    bom = inst.get("bommanahalli", {"online": 0, "offline": 0, "offline_pf": 0, "total": 0})
    east = inst.get("east", {"online": 0, "offline": 0, "offline_pf": 0, "total": 0})
    comb = inst.get("combined", {"online": 0, "offline": 0, "offline_pf": 0, "total": 0})
    issues = inst.get("issues", {})
    perf_score = inst.get("performance_score", 0.0)
    kpi_score = inst.get("kpi_score", 0.0)
    penalty_pts = inst.get("penalty_points", 0)

    print("\n" + "="*110)
    print(f"⚡ {proj_name.upper()} PANEL TELEMETRY REPORT | PANEL PERFORMANCE SCORE: {perf_score}% | KPI SCORE (excl. PF): {kpi_score}%")
    print(" (Performance Formula: (Online + Offline PF Today) / Total Panels)")
    print("="*110)
    print(f"{'Region / Zone':<26} | {'Online':<8} | {'Offline':<8} | {'PF (Today)':<12} | {'PF (Prior)':<12} | {'Total Panels':<10}")
    print("-" * 110)
    if regions:
        for r in regions:
            r_name = r.get("name", "Region")
            print(f"{r_name:<26} | {r.get('online', 0):<8} | {r.get('offline', 0):<8} | {r.get('offline_pf_today', 0):<12} | {r.get('offline_pf_prior', 0):<12} | {r.get('total', 0):<10}")
    elif proj_name.upper() != "BBMP":
        print(f"{proj_name:<26} | {comb.get('online', 0):<8} | {comb.get('offline', 0):<8} | {comb.get('offline_pf_today', 0):<12} | {comb.get('offline_pf_prior', 0):<12} | {comb.get('total', 0):<10}")
    else:
        print(f"{'Bommanahalli':<26} | {bom.get('online', 0):<8} | {bom.get('offline', 0):<8} | {bom.get('offline_pf_today', 0):<12} | {bom.get('offline_pf_prior', 0):<12} | {bom.get('total', 0):<10}")
        print(f"{'EAST':<26} | {east.get('online', 0):<8} | {east.get('offline', 0):<8} | {east.get('offline_pf_today', 0):<12} | {east.get('offline_pf_prior', 0):<12} | {east.get('total', 0):<10}")
    print("-" * 110)
    print(f"{'COMBINED / TOTAL':<26} | {comb.get('online', 0):<8} | {comb.get('offline', 0):<8} | {comb.get('offline_pf_today', 0):<12} | {comb.get('offline_pf_prior', 0):<12} | {comb.get('total', 0):<10}")
    print("="*110)

    print("\n" + "="*95)
    print(f"⚠️  PANEL ISSUES BREAKDOWN (O&M DASHBOARD) | TOTAL PENALTY POINTS: {penalty_pts}")
    print("="*95)
    print(f"{'Input Issues':<28} | {'Output Issues':<32} | {'Other Issues':<28}")
    print("-" * 95)
    print(f"{'Low Voltage:':<20} {issues.get('low_voltage', 0):<7} | {'High Current / Power Theft:':<26} {issues.get('high_current', 0):<5} | {'Relay Failure:':<20} {issues.get('relay_failure', 0):<7}")
    print(f"{'High Voltage:':<20} {issues.get('high_voltage', 0):<7} | {'Low Current:':<26} {issues.get('low_current', 0):<5} | {'MeterComm Failure:':<20} {issues.get('meter_comm_failure', 0):<7}")
    print(f"{'':<28} | {'MCB Trip:':<26} {issues.get('mcb_trip', 0):<5} | {'':<28}")
    print("-" * 95)
    print(f"🚨 LONG-TERM OFFLINE BREAKDOWN (> 7 DAYS):")
    print(f"  • Offline Panels (> 7 Days):    {inst.get('offline_gt_7_days', issues.get('offline_gt_7_days', 0))}")
    print(f"  • Offline PF Panels (> 7 Days): {inst.get('offline_pf_gt_7_days', issues.get('offline_pf_gt_7_days', 0))}")
    print("="*95 + "\n")


import csv

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_panel_issues_to_csv(data: dict, output_csv: str = "panel_issues_details.csv") -> str:
    """Exports granular breakdown of panels with specific dashboard issues to CSV."""
    inst = data.get("installation_report", {})
    affected_panels = inst.get("affected_panels", [])
    csv_path = Path(__file__).resolve().parent / output_csv
    
    fieldnames = ["Device ID", "Panel Name", "Panel Label", "Region", "Zone Name", "Ward Name", "Status", "Last Received Data Date", "Days Offline", "Active Issues", "Lat / Lon"]
    
    exported_count = 0
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(fieldnames)
        for p in affected_panels:
            active_issues = p.get("active_issues", [])
            if not active_issues:
                continue
            writer.writerow([
                p.get("id", ""),
                p.get("name", ""),
                p.get("label", ""),
                p.get("region", ""),
                p.get("zone", ""),
                p.get("ward", ""),
                p.get("status", ""),
                p.get("last_received_date", "-"),
                p.get("days_offline", "-"),
                ", ".join(active_issues),
                p.get("lat_lon", "-")
            ])
            exported_count += 1
            
    logger.info(f"Exported detailed panel issues ({exported_count} panels) to {csv_path}")
    return str(csv_path)


def export_panel_issues_to_excel(data: dict, output_excel: str = "panel_issues_details.xlsx", project_name: str = "BBMP") -> str:
    """Exports granular breakdown of panels into Excel tabs (regional tabs for BBMP, single sheet for 5B Innovation)."""
    inst = data.get("installation_report", {})
    affected_panels = inst.get("affected_panels", [])
    excel_path = Path(__file__).resolve().parent / output_excel

    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Header Style (Dark Blue Fill with Bold White Text)
    header_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Thin Grid Border
    thin_side = Side(style="thin", color="CBD5E0")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Data Style
    data_font = Font(name="Calibri", size=10)
    data_alignment = Alignment(horizontal="left", vertical="center")

    fieldnames = ["Panel Name", "Panel Label", "Region", "Zone Name", "Ward Name", "Status", "Last Received Data Date", "Days Offline", "Active Issues", "Lat / Lon"]

    link_font = Font(name="Calibri", size=10, color="0000FF", underline="single")

    def create_sheet(title, panels):
        ws = wb.create_sheet(title=title)
        ws.append(fieldnames)

        # Format Headers
        for col_idx, header_cell in enumerate(ws[1], start=1):
            header_cell.fill = header_fill
            header_cell.font = header_font
            header_cell.alignment = header_alignment
            header_cell.border = thin_border
        ws.row_dimensions[1].height = 25

        row_idx = 2
        for p in panels:
            lat_lon_val = p.get("lat_lon", "-")
            ws.append([
                p.get("name", ""),
                p.get("label", ""),
                p.get("region", ""),
                p.get("zone", ""),
                p.get("ward", ""),
                p.get("status", ""),
                p.get("last_received_date", "-"),
                p.get("days_offline", "-"),
                ", ".join(p.get("active_issues", [])),
                lat_lon_val
            ])
            for col_idx, cell in enumerate(ws[row_idx], start=1):
                cell.font = data_font
                cell.border = thin_border
                if col_idx in (3, 6, 7, 8, 10): # Region, Status, Last Received Data Date, Days Offline & Lat / Lon centered
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = data_alignment

            # Format Lat / Lon cell as clickable Google Maps hyperlink if valid coordinates
            if lat_lon_val and lat_lon_val != "-" and "," in str(lat_lon_val):
                coords = [c.strip() for c in str(lat_lon_val).split(",")]
                if len(coords) == 2:
                    try:
                        lat_f, lon_f = float(coords[0]), float(coords[1])
                        maps_url = f"https://www.google.com/maps?q={lat_f},{lon_f}"
                        lat_lon_cell = ws.cell(row=row_idx, column=10)
                        lat_lon_cell.hyperlink = maps_url
                        lat_lon_cell.font = link_font
                    except ValueError:
                        pass

            ws.row_dimensions[row_idx].height = 20
            row_idx += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    is_bbmp = (project_name.upper() == "BBMP")

    if is_bbmp:
        # Filter panels by region for BBMP
        bom_panels = []
        east_panels = []
        other_panels = []

        for p in affected_panels:
            if not p.get("active_issues"):
                continue
            reg_upper = str(p.get("region", "")).upper()
            if "BOMMANAHALLI" in reg_upper or "BOMMANAHALI" in reg_upper or "BMH" in reg_upper:
                bom_panels.append(p)
            elif "EAST" in reg_upper:
                east_panels.append(p)
            else:
                other_panels.append(p)

        create_sheet("Bommanahalli", bom_panels)
        create_sheet("East", east_panels)
        if other_panels:
            create_sheet("Other Regions", other_panels)
        logger.info(f"Exported BBMP detailed panel issues (Bommanahalli: {len(bom_panels)}, East: {len(east_panels)}) to Excel: {excel_path}")
    else:
        # 5B Innovation or non-BBMP project - Single dedicated sheet without Bommanahalli / East tabs
        active_5b_panels = [p for p in affected_panels if p.get("active_issues")]
        sheet_title = f"{project_name} Panel Issues" if len(f"{project_name} Panel Issues") <= 30 else "Panel Issues Details"
        create_sheet(sheet_title, active_5b_panels)
        logger.info(f"Exported {project_name} detailed panel issues ({len(active_5b_panels)} panels) to single sheet '{sheet_title}' Excel: {excel_path}")

    wb.save(excel_path)
    return str(excel_path)



def list_customers_cmd():
    """Authenticates with ThingsBoard and prints all available Customers and their IDs."""
    logger.info("Connecting to ThingsBoard to list available Customers...")
    tb_client = ThingsBoardClient()
    if tb_client.login():
        customers = tb_client.get_tenant_customers()
        print("\n" + "="*85)
        print("🏢 THINGSBOARD CUSTOMERS LIST")
        print("="*85)
        print(f"{'Customer Title / Name':<45} | {'Customer ID (UUID)':<36}")
        print("-" * 85)
        if not customers:
            print("No customers found or account lacks customer list permission.")
        for cust in customers:
            title = cust.get("title", cust.get("name", "Unknown"))
            cid = cust.get("id", {}).get("id", "-")
            print(f"{title:<45} | {cid:<36}")
        print("="*85 + "\n")
    else:
        logger.error("Failed to authenticate with ThingsBoard API.")


def run_pipeline(
    dry_run: bool = False,
    use_mock: bool = False,
    print_cli: bool = False,
    output_file: str = "report_preview.html",
    override_recipients: Optional[List[str]] = None,
    override_bcc: Optional[List[str]] = None,
    project_name: Optional[str] = None,
    customer_id: Optional[str] = None,
    subject_prefix: Optional[str] = None,
    thread_id: Optional[str] = None
):
    """
    Main job function: Fetches data, displays terminal summary, generates HTML report, exports issue details Excel, and sends email or saves preview.
    """
    raw_proj = str(project_name or Config.PROJECT_NAME or "BBMP").strip()
    if "5B" in raw_proj.upper() or "INNOVATION" in raw_proj.upper():
        proj_name = "5B Innovation"
    else:
        proj_name = "BBMP"

    logger.info(f"=== Starting Panel Report Job for Project: {proj_name} ===")

    if override_recipients:
        recipients_to_use = override_recipients
    elif proj_name != "BBMP":
        recipients_to_use = Config.RECIPIENT_EMAILS_5B or Config.RECIPIENT_EMAILS
    else:
        recipients_to_use = Config.RECIPIENT_EMAILS

    if override_bcc is not None:
        bcc_to_use = override_bcc
    elif override_recipients:
        # If recipient address is explicitly overridden (e.g. via --to),
        # do NOT default to the config BCC list unless --bcc was explicitly specified.
        bcc_to_use = []
    elif proj_name != "BBMP":
        bcc_to_use = Config.BCC_EMAILS_5B
    else:
        bcc_to_use = Config.BCC_EMAILS

    if proj_name != "BBMP":
        subj_prefix = subject_prefix or Config.EMAIL_SUBJECT_PREFIX_5B
        t_id = thread_id or Config.EMAIL_THREAD_ID_5B
        enable_threading = Config.ENABLE_EMAIL_THREADING_5B
    else:
        subj_prefix = subject_prefix or Config.EMAIL_SUBJECT_PREFIX
        t_id = thread_id or Config.EMAIL_THREAD_ID
        enable_threading = Config.ENABLE_EMAIL_THREADING

    # 1. Fetch Data
    if use_mock:
        logger.info(f"Using mock panel data for {proj_name} report generation.")
        data = get_mock_panel_data(proj_name)
    else:
        tb_missing = Config.validate(check_smtp=False, check_tb=True)
        if tb_missing:
            logger.error(f"Missing ThingsBoard credentials ({', '.join(tb_missing)}). Cannot proceed with live report execution.")
            raise RuntimeError(f"Missing ThingsBoard configuration: {', '.join(tb_missing)}")
        
        tb_client = ThingsBoardClient()
        if tb_client.login():
            logger.info(f"Fetching Panel Installation Report for Customer/Project: {proj_name}...")
            inst_report = tb_client.fetch_panel_installation_report(customer_id=customer_id, project_name=proj_name)
            if not inst_report or inst_report.get("total_panels", 0) == 0:
                logger.error(f"Data fetching failed for project '{proj_name}' (0 panels retrieved). Aborting email report sending to prevent sending empty report.")
                raise RuntimeError(f"Data fetching failed for project '{proj_name}': 0 panels retrieved.")

            alarms = tb_client.get_active_alarms()
            data = {
                "installation_report": inst_report,
                "alarms": alarms,
                "summary": {
                    "total_devices": inst_report.get("total_panels", 0),
                    "online_devices": inst_report.get("combined", {}).get("online", 0),
                    "offline_devices": inst_report.get("combined", {}).get("offline", 0),
                    "active_alarms": len(alarms)
                }
            }
        else:
            logger.error(f"Failed to authenticate with ThingsBoard API for project '{proj_name}' after retries. Aborting email report sending to prevent sending empty report.")
            raise ConnectionError(f"ThingsBoard authentication failed for project '{proj_name}'. Execution aborted.")

    # Always print terminal summary table
    print_terminal_summary(data)

    # Export detailed issues to Excel (.xlsx) with Google Maps hyperlinks
    excel_filename = f"{proj_name.lower().replace(' ', '_')}_panel_issues_details.xlsx" if proj_name != "BBMP" else "panel_issues_details.xlsx"
    excel_file_path = export_panel_issues_to_excel(data, output_excel=excel_filename, project_name=proj_name)

    # 2. Generate HTML Report
    logger.info("Generating HTML report...")
    html_report = generate_html_report(data)

    # 3. Always save a local preview file
    out_file = output_file
    if output_file == "report_preview.html" and proj_name != "BBMP":
        out_file = f"{proj_name.lower().replace(' ', '_')}_report_preview.html"
        
    preview_path = Path(__file__).resolve().parent / out_file
    with open(preview_path, "w", encoding="utf-8") as f:
        f.write(html_report)
    logger.info(f"Report saved to local file: {preview_path}")

    # 4. Handle Email Sending or Dry Run / CLI mode
    if dry_run or print_cli:
        logger.info(f"DRY-RUN / PRINT Mode: Email sending skipped. You can open '{out_file}' in your browser to view the formatted report.")
        return True

    # Validate SMTP configuration before sending
    smtp_missing = Config.validate(check_smtp=True, check_tb=False)
    if smtp_missing and not override_recipients:
        logger.error(f"Cannot send email. Missing SMTP configuration/secrets: {', '.join(smtp_missing)}")
        logger.error("Please add SMTP_USERNAME, SMTP_PASSWORD, and RECIPIENT_EMAILS into GitHub Repository Secrets or .env file.")
        sys.exit(1)

    if not recipients_to_use:
        logger.error("No recipient email address available. Use --to your_email@domain.com or set RECIPIENT_EMAILS in .env")
        sys.exit(1)

    mailer = EmailSender(
        enable_threading=enable_threading,
        thread_id=t_id
    )
    subject = f"{subj_prefix} Telemetry Status Report"
    
    # Attach Excel sheet only
    attachments_to_send = [excel_file_path]
    inline_imgs = None

    logger.info(f"Sending email report to: {', '.join(recipients_to_use)}" + (f" (BCC: {', '.join(bcc_to_use)})" if bcc_to_use else ""))
    success = mailer.send_email(
        recipients=recipients_to_use,
        subject=subject,
        html_content=html_report,
        attachment_paths=attachments_to_send,
        inline_images=inline_imgs,
        bcc_recipients=bcc_to_use,
        enable_threading=enable_threading,
        thread_id=t_id
    )

    if success:
        logger.info(f"{proj_name} Panel Report job completed successfully!")
    else:
        logger.error(f"{proj_name} Panel Report job encountered email sending errors.")
        sys.exit(1)

    return success


def inspect_project_regions_cmd(project_name: Optional[str] = None, customer_id: Optional[str] = None):
    """Fetches and displays all unique regions, zones, wards, and panel counts for specified project/customer."""
    proj_name = project_name or Config.PROJECT_NAME or "5B Innovations"
    if "5B" in proj_name.upper() or "INNOVATION" in proj_name.upper():
        proj_name = "5B Innovations"

    logger.info(f"Inspecting regions and zones for project: {proj_name}...")
    tb_client = ThingsBoardClient()
    if tb_client.login():
        target_cust_id = customer_id
        if not target_cust_id:
            if "5B" in proj_name.upper() or "INNOVATION" in proj_name.upper():
                target_cust_id = Config.TB_CUSTOMER_ID_5B
                if not target_cust_id:
                    matched_cust = tb_client.find_customer_by_name("5B") or tb_client.find_customer_by_name("Innovation")
                    if matched_cust:
                        target_cust_id = matched_cust.get("id", {}).get("id")
            if not target_cust_id:
                target_cust_id = Config.TB_CUSTOMER_ID

        devices = []
        page = 0
        while True:
            url = f"{tb_client.host}/api/customer/{target_cust_id}/deviceInfos?pageSize=1000&page={page}"
            res = tb_client.session.get(url, timeout=30)
            if res.status_code != 200:
                url = f"{tb_client.host}/api/customer/{target_cust_id}/devices?pageSize=1000&page={page}"
                res = tb_client.session.get(url, timeout=30)
            if res.status_code == 200:
                data = res.json()
                devs = data.get("data", [])
                devices.extend(devs)
                if not data.get("hasNext", False) or len(devs) == 0:
                    break
                page += 1
            else:
                break
        
        print("\n" + "="*90)
        print(f"📍 {proj_name.upper()} REGIONS, ZONES & PANEL COUNT BREAKDOWN (Total: {len(devices)} panels)")
        print("="*90)

        regions_counter = {}
        zones_counter = {}

        def check_attr(dev):
            dev_id = dev.get("id", {}).get("id")
            reg = "NOT_SET"
            zone = "NOT_SET"
            ward = "NOT_SET"
            try:
                res_attr = tb_client.session.get(f"{tb_client.host}/api/plugins/telemetry/DEVICE/{dev_id}/values/attributes?keys=region,zoneName,wardName", timeout=5)
                if res_attr.status_code == 200:
                    for item in res_attr.json():
                        k = item.get("key")
                        v = item.get("value")
                        if k == "region" and v:
                            reg = str(v)
                        elif k == "zoneName" and v:
                            zone = str(v)
                        elif k == "wardName" and v:
                            ward = str(v)
            except Exception:
                pass
            return reg, zone, ward

        from concurrent.futures import ThreadPoolExecutor, as_completed
        with ThreadPoolExecutor(max_workers=30) as executor:
            futures = [executor.submit(check_attr, dev) for dev in devices]
            for future in as_completed(futures):
                reg, zone, ward = future.result()
                regions_counter[reg] = regions_counter.get(reg, 0) + 1
                zones_counter[zone] = zones_counter.get(zone, 0) + 1

        print(f"\n--- REGIONS FOUND ({len(regions_counter)}) ---")
        print(f"{'Region Attribute Value':<45} | {'Panel Count':<15}")
        print("-" * 65)
        for r_name, count in sorted(regions_counter.items(), key=lambda x: x[1], reverse=True):
            print(f"{r_name:<45} | {count:<15}")

        print(f"\n--- ZONES FOUND ({len(zones_counter)}) ---")
        print(f"{'Zone Attribute Value':<45} | {'Panel Count':<15}")
        print("-" * 65)
        for z_name, count in sorted(zones_counter.items(), key=lambda x: x[1], reverse=True):
            print(f"{z_name:<45} | {count:<15}")
        print("="*90 + "\n")
    else:
        logger.error("Failed to authenticate with ThingsBoard API.")


def main():
    parser = argparse.ArgumentParser(description="BBMP & 5B Innovation ThingsBoard Panel Monitoring & Email Report Tool")
    parser.add_argument("--project", type=str, default=None, help="Project name (e.g. 'bbmp', '5b_innovation')")
    parser.add_argument("--customer-id", type=str, default=None, help="Target ThingsBoard Customer ID (UUID)")
    parser.add_argument("--list-customers", action="store_true", help="List all available Customers and Customer IDs on ThingsBoard account")
    parser.add_argument("--inspect-regions", action="store_true", help="Inspect all unique regions and zones for specified project")
    parser.add_argument("--print", "--cli", action="store_true", help="Fetch data and print summary directly in terminal without sending email")
    parser.add_argument("--dry-run", action="store_true", help="Fetch data and generate report preview without sending emails")
    parser.add_argument("--send-now", action="store_true", help="Fetch data and send email immediately")
    parser.add_argument("--mock", action="store_true", help="Use mock telemetry data for testing")
    parser.add_argument("--to", "--recipient", type=str, help="Override recipient email address(es), comma-separated")
    parser.add_argument("--bcc", type=str, help="Override BCC recipient email address(es), comma-separated")
    parser.add_argument("--output", type=str, default="report_preview.html", help="Local preview output filename")

    args = parser.parse_args()

    if args.list_customers:
        list_customers_cmd()
        return

    if args.inspect_regions:
        inspect_project_regions_cmd(project_name=args.project, customer_id=args.customer_id)
        return

    override_recipients = None
    if args.to:
        override_recipients = [r.strip() for r in args.to.split(",") if r.strip()]

    override_bcc = None
    if args.bcc is not None:
        override_bcc = [r.strip() for r in args.bcc.split(",") if r.strip() and r.lower() != "none"]

    # If --mock is used without --send-now, default dry_run to True so mock data doesn't send email
    dry_run_mode = args.dry_run
    if args.mock and not args.send_now:
        dry_run_mode = True

    projects_to_run = []
    if args.project:
        p_arg = args.project.strip().lower()
        if p_arg in ("all", "both"):
            projects_to_run = ["BBMP", "5B Innovation"]
        elif "5b" in p_arg or "innovation" in p_arg:
            projects_to_run = ["5B Innovation"]
        else:
            projects_to_run = [args.project]
    else:
        env_proj = (Config.PROJECT_NAME or "BBMP").strip()
        if env_proj.lower() in ("all", "both"):
            projects_to_run = ["BBMP", "5B Innovation"]
        elif "5b" in env_proj.lower() or "innovation" in env_proj.lower():
            projects_to_run = ["5B Innovation"]
        elif args.send_now:
            # When --send-now is triggered without explicit project flag, send separate emails for both projects
            projects_to_run = ["BBMP", "5B Innovation"]
        else:
            projects_to_run = ["BBMP"]

    logger.info(f"Projects queued for report execution: {', '.join(projects_to_run)}")

    results = []
    for proj in projects_to_run:
        out_f = args.output
        if out_f == "report_preview.html" and proj != "BBMP":
            out_f = f"{proj.lower().replace(' ', '_')}_report_preview.html"

        res = run_pipeline(
            dry_run=dry_run_mode,
            use_mock=args.mock,
            print_cli=args.print or (len(projects_to_run) == 1 and not (args.send_now or args.dry_run)),
            output_file=out_f,
            override_recipients=override_recipients,
            override_bcc=override_bcc,
            project_name=proj,
            customer_id=args.customer_id if len(projects_to_run) == 1 else None
        )
        results.append(res)

    if not all(results):
        sys.exit(1)


if __name__ == "__main__":
    main()

